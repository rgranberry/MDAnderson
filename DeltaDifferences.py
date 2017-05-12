import openpyxl
import Utils

"""
This function takes two separate Excel workbooks and finds the delta of every value in each cell and puts
those values into a new sheet.

Hardcoded values:
    Name of column by which to differentiate patients
    Starting row
    Starting column
    Starting column with numerical data
"""


def delta_differences(sheet_name, sheet_name2, new_sheet_name, sheet_names):

    # open the appropriate document and access the first workbook_1 in it
    doc = openpyxl.load_workbook(sheet_name)
    doc2 = openpyxl.load_workbook(sheet_name2)
    doc3 = openpyxl.load_workbook(new_sheet_name)

    workbook_1 = doc.get_sheet_by_name(sheet_names[0])
    workbook_2 = doc2.get_sheet_by_name(sheet_names[0])

    num_cols1 = workbook_1.max_column
    num_rows1 = workbook_1.max_row

    num_cols2 = workbook_2.max_column
    num_rows2 = workbook_2.max_row

    # find the column that contains the MRN values
    mrn_col = Utils.find_index_column(workbook_1, ' Image', 1)

    # Find the names of the patients that are only in one sheet (they should be ignored in the final document)
    pre_names = []
    print ("num_rows2: ", num_rows1)
    for row in range(5, num_rows1 + 1):
        name = workbook_1[mrn_col + str(row)].value.lower().split('^')[0]
        print("name: ", name)
        pre_names.insert(0, name)

    print("pre_names: ", pre_names)

    # Create the appropriate sheets in doc3
    Utils.create_remove_sheets(doc3, sheet_names)

    doc3.save(new_sheet_name)

    print "Shouldn't find kelley, lewis, morris, perez"
    for idx in range(5, num_rows2 + 1):

        patient_name = workbook_2[mrn_col + str(idx)].value.lower().split('^')[0]

        if patient_name not in pre_names:
            print ("Patient name not found")
            print ("Patient name: ", patient_name)
            continue
        else:
            # Identify the row from which to get the item in sheet_1
            sheet2_row = str(idx)

            # Identify the row from which to get the item in sheet_2
            for row in range(5, num_rows1):
                if workbook_1[mrn_col + str(row)].value.lower().split('^')[0] == patient_name:
                    sheet1_row = str(row)
                    continue

            col_count = 1
            first_letter = 65
            third_letter = 64
            use_third = False

            for col in range(1, num_cols2 + 1):

                ascii_data = Utils.calculate_ascii(col, col_count, use_third, first_letter, third_letter)
                ascii_col = ascii_data[0]
                col_count = ascii_data[1]
                use_third = ascii_data[2]
                first_letter = ascii_data[3]
                third_letter = ascii_data[4]

                # getString is the cell from which to grab the current information
                get_string1 = ascii_col + sheet1_row
                get_string2 = ascii_col + sheet2_row

                for sheet in sheet_names:
                    # print ("sheet: ", sheet)
                    # iterate over every sheet and add the correct data to it
                    sheet_1 = doc.get_sheet_by_name(sheet)
                    sheet_2 = doc2.get_sheet_by_name(sheet)
                    sheet_3 = doc3.get_sheet_by_name(sheet)

                    val1 = sheet_1[get_string1].value
                    val2 = sheet_2[get_string2].value

                    sheet_3[ascii_col + '1'].value = sheet_1[ascii_col + '1'].value

                    if col > 5:
                        if val1 == 0:
                            new_val = 0
                        elif (val2 is None) or (val1 is None):
                            new_val = None
                        else:
                            new_val = 1.0 * (val2 - val1) / val1
                            sheet_3[get_string2].value = new_val
                    else:
                        sheet_3[get_string2].value = sheet_2[get_string2].value

    doc3.remove_sheet(doc3.get_sheet_by_name('Sheet1'))
    # save all of the documents at the end
    doc3.save(new_sheet_name)


# Sheet_names should be the names of each sheet you want converted
# The first item in run_excel should be the name of the file - run_excel('excel_document.xlsx', sheet_names)
sheet_names = ['Mandible', 'Cropped', 'ORN']
delta_differences('Pre-RT-Final.xlsx', 'Post-RT-Final.xlsx', 'Delta Results.xlsx', sheet_names)