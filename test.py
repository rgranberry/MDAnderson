import openpyxl
# import cProfile


def run_excel(sheet_name, sheet_name2, new_sheet_name, sheet_names):

    use_third = False

    # open the appropriate documents
    doc = openpyxl.load_workbook(sheet_name)
    doc2 = openpyxl.load_workbook(sheet_name2)
    doc3 = openpyxl.load_workbook(new_sheet_name)

    # use the first sheet in documents 1 and 2 for data (because all other sheets are the same format)
    workbook_1 = doc.get_sheet_by_name(sheet_names[0])
    workbook_2 = doc2.get_sheet_by_name(sheet_names[0])

    num_cols = workbook_1.max_column
    num_rows = workbook_1.max_row

    col_count = 1
    first_letter = 65
    third_letter = 64

    # iterate over all of the columns in the document

    # find the column that contains the MRN/index values
    for ind in range(1, 26):
        if workbook_1[chr(ind + 64) + str(1)].value == 'Index':
            mrn_col = chr(64 + ind)
            break

    post_patients = []

    # Find all of the names that are only in doc2 not in doc (because you don't want extra ones in doc3)
    for row in range(2, num_rows):
        post_patients.insert(0, workbook_2[mrn_col + str(row)].value)

    print("Sheet names: ", doc3.sheetnames)

    # doc3.create_sheet(title="hello")
    # doc3.save("Workbook3.xlsx")
    # Create the appropriate sheets in doc3 (removing them if any previous ones have been created)
    for sheet in sheet_names:

        if sheet in doc3.sheetnames and sheet is not sheet_names[0]:
            print("Removed sheet")
            remove_sheet = doc3.get_sheet_by_name(sheet)
            doc3.remove_sheet(remove_sheet)
        print("Created ", sheet)
        doc3.create_sheet(sheet)
    doc3.save(new_sheet_name)

    print("SheetNames: ", doc3.sheetnames)
    for idx in range(2, num_rows + 1):

        print("Iterating through rows")
        patient_name = workbook_1[mrn_col + str(idx)].value

        if patient_name not in post_patients:
            print("Patient name not in doc2")
            continue
        else:
            # Sheet1Row is the row from which you should get values in sheet_1
            sheet1_row = str(idx)

            # Sheet2Row is the row from which you should get values in sheet_2
            for row in range(2, num_rows):
                print("Patient name: ", patient_name)
                print("Sheet-2 value: ", workbook_2['C' + str(row)].value)
                if workbook_2[mrn_col + str(row)].value == patient_name:
                    sheet2_row = str(row)

            for col in range(1, num_cols + 1):
                print("Iterating through columns")

                # Determine the letter code corresponding to the column in the excel sheet
                if col <= 26:
                    # if it's under 26 columns, just use a single letter
                    ascii_col = chr(col + 64)
                elif use_third:
                    if col_count > 26:
                        # first_letter describes the coordinate of what the first letter should be -
                        # every 26 iterations, it increases by one to switch the first letter up by one
                        first_letter += 1
                        # col_count keeps track of what column you're at in the current first_letter iteration
                        col_count = 1
                    if first_letter > 90:
                        third_letter += 1
                        first_letter = 65
                    ascii_col = chr(third_letter) + chr(first_letter) + chr((col_count + 64))

                    col_count += 1
                else:
                    # if it's over 26 columns, you have to calculate two different letters
                    if col_count > 26:
                        # first_letter describes the coordinate of what the first letter should be -
                        # every 26 iterations, it increases by one to switch the first letter up by one
                        first_letter += 1
                        # col_count keeps track of what column you're at in the current first_letter iteration
                        col_count = 1

                    ascii_col = chr(first_letter) + chr((col_count + 64))

                    if ascii_col == 'ZZ':
                        use_third = True

                    col_count += 1

                # getString is the cell from which to grab the current information from either doc or doc2
                get_string1 = ascii_col + sheet1_row
                get_string2 = ascii_col + sheet2_row

                # Iterate through each of the sheets and calculate the new value that needs to be put into that sheet
                # in doc3
                for sheet in sheet_names:
                    sheet_1 = doc.get_sheet_by_name(sheet)
                    sheet_2 = doc2.get_sheet_by_name(sheet)
                    sheet_3 = doc3.get_sheet_by_name(sheet)

                    sheet_3[ascii_col + '1'].value = sheet_1[ascii_col + '1'].value

                    if col > 2:
                        new_val = 1.0 * (sheet_2[get_string2].value - sheet_1[get_string1].value) / sheet_1[get_string1].value
                        sheet_3[get_string2].value = new_val

                    else:

                        sheet_3[get_string2].value = sheet_2[get_string2].value

    # save all of the documents at the end
    doc3.save(new_sheet_name)


# Sheet_names should be the names of each sheet you want converted
# The first item in run_excel should be the name of the file - run_excel('excel_document.xlsx', sheet_names)
sheets = ['Sheet1', 'Sheet2']
run_excel('Workbook1.xlsx', 'Workbook2.xlsx', 'Workbook4.xlsx', sheets)