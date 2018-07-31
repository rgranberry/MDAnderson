import openpyxl
import re
import tkinter as tk
from tkinter import filedialog
import os

def bottle_data(sheet_names, start_month, start_day, end_month, end_day):
    # open the appropriate document and access the correct sheets

    root = tk.Tk()
    root.withdraw()

    pre_doc_name1 = filedialog.askopenfilename()
    split_doc1 = pre_doc_name1.split('/')
    doc_name1 = split_doc1[-1]
    os.chdir("/".join(split_doc1[:len(split_doc1)-1]))
    doc = openpyxl.load_workbook(doc_name1)

    pre_doc_name2 = filedialog.askopenfilename()
    split_doc2 = pre_doc_name2.split('/')
    doc_name2 = split_doc2[-1]
    os.chdir("/".join(split_doc2[:len(split_doc2)-1]))
    doc2 = openpyxl.load_workbook(doc_name2, keep_vba = False)

    sheet_1 = doc[sheet_names[0]]
    sheet_2 = doc2[sheet_names[1]]

    start_date_found = False

    weight_values = []

    day = None
    month = None
    i = 0

    for col in sheet_1.iter_cols():

        day_values = []
        i += 1

        # Iterate through the columns to find the date for each set
        if col[1].value is not None:
            day = str(col[1].value.day)
            month = str(col[1].value.month)
            date = month + "/" + day

            # sequence is whether the set is a "before" or "after" value
            sequence = str(sheet_1.cell(row=1, column=i).value)

        if day == start_day and month == start_month:
            start_date_found = True
        if day == end_day and month == end_month:
            break
        if start_date_found is True:
            # liquid is whether the set is alcohol or water
            liquid = str(sheet_1.cell(row=4, column=i).value)

            # add all of the distinguishing factors (date, liquid, and before/after) to the set for later use
            day_values.append(date)
            day_values.append(liquid)
            day_values.append(sequence)

            # iterates through every row in the set
            for row in range(5, sheet_1.max_row):
                cell_value = sheet_1.cell(row=row, column=i).value
                if cell_value is None:
                    # will stop running the iterations as soon as it doesn't find any more values
                    break
                # removes the grams from the bottle weight value
                bottle_weight_corrected = re.findall(r"[-+]?\d*\.\d+|\d+", str(cell_value))
                # appends each bottle weight to the set for that day
                day_values.append(float(bottle_weight_corrected[0]))

        # appends the single set values into the larger overall set
        if len(day_values) != 0:
            weight_values.append(day_values)

    # enters all of the data calculated above into the second spreadsheet

    for i in range(0, len(weight_values)):
        set = weight_values[i]
        if set[2] == "After":
            # Calculates the spill values and adds the difference into the after values
            spill_before = weight_values[i-2][-1]
            spill_after = set[-1]
            spill_difference = spill_before - spill_after
            for value in range(3, len(set)):
                set[value] = set[value] + spill_difference

    for set in weight_values:
        for row_num in range(2, sheet_2.max_row):
            cell_name = "A" + str(row_num)
            # iterates through every set
            if set[0] in str(sheet_2[cell_name].value):
                if "20%" in set[1]:
                    # if the value for "liquid" was EtOH 20%, adds the values to the second spreadsheet in cols C and D
                    add_values(set, row_num, 'C', 'D', sheet_2)
                else:
                    # if the value for "liquid" was H2O, adds the values to the second spreadsheet in cols G and H
                    add_values(set, row_num, 'G', 'H', sheet_2)

    # save the changes to the second spreadsheet
    doc2.save(doc_name2)


def add_values(set, row_num, letter1, letter2, sheet_2):
    # adds the values calculated into the second spreadsheet
    if set[2] == "Before":
        idx = row_num + 1
        for i in range(3, len(set) - 1):
            sheet_2[letter1 + str(idx)].value = set[i]
            idx += 1
    else:
        idx = row_num + 1
        for i in range(3, len(set) - 1):
            sheet_2[letter2 + str(idx)].value = set[i]
            idx += 1


bottle_data(["Bottle Weights", "Drinking Data"], "7", "18", "7", "20")
