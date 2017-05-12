import openpyxl
import Utils

"""
This function takes one excel workbook and reorganizes it so that a new sheet is created where each patient
gets their own row with all of their values (condenses a sheet with many of the same MRNs at different points
in time and puts them all in one).

Hardcoded values:
    Name of column by which to differentiate patients
    Starting row
    Starting column
"""


def parotid(sheet_name, sheet1):

    # open the appropriate document and access the first sheet in it
    doc = openpyxl.load_workbook(sheet_name)

    sheet = doc.get_sheet_by_name(sheet1)
    doc.save(sheet_name)

    num_cols = sheet.max_column
    num_rows = sheet.max_row

    col_count = 1
    first_letter = 65
    use_third = False
    third_letter = 64

    # find the column that contains the MRN values
    mrn_col = Utils.find_index_column(sheet, ' MRN', 2)

    # iterate over all of the columns in the document
    for col in range(1, num_cols + 1):

        ascii_data = Utils.calculate_ascii(col, col_count, use_third, first_letter, third_letter)
        ascii_col = ascii_data[0]
        col_count = ascii_data[1]
        use_third = ascii_data[2]
        first_letter = ascii_data[3]
        third_letter = ascii_data[4]
        # print("ASCII COL: ", ascii_col)

        # if a sheet with the name you're trying to add already exists, delete it and then re-add the new one
        new_sheet_name = str(sheet[ascii_col + str(1)].value)

        Utils.create_remove_sheets(doc, [new_sheet_name])

        sheet2 = doc.get_sheet_by_name(new_sheet_name)

        # row_count keeps track of which row you're in in the new sheet that you're adding to (how many of each MRN
        #  there are)
        row_count = 1
        # new_col keeps track of the current overall column you're in for each new MRN
        new_col = 2
        # count keeps track of the second letter in the column past 25 (equivalent to col_count)
        count = 0
        # keeps track of which letter you're in (equivalent of first_letter)
        first_col = 65

        for idx in range(2, num_rows + 1):

            # define the MRN columns - which ones to separate out patients by
            mrn1 = mrn_col + str(idx)
            mrn2 = mrn_col + str(idx - 1)

            # get_string is the cell from which to grab the current information
            get_string = ascii_col + str(idx)

            # do this every time you encounter a new MRN
            if sheet[mrn1].value != sheet[mrn2].value:
                # increase the row_count (added a new patient)
                row_count += 1
                # reset the values of new_col, count, and first_col for the new patient
                new_col = 2
                count = 0
                first_col = 65
                # make the first cell in each patient's row be their MRN value
                sheet2['A' + str(row_count)].value = sheet[mrn1].value

            if new_col > 26:
                # if you've surpassed 26 cells, calculate the double letter value for their column
                count += 1
                if count > 26:
                    first_col += 1
                    count = 1
                new_string = chr(first_col) + chr(count + 64) + str(row_count)
            else:
                new_string = chr(new_col + 64) + str(row_count)

            # make the value at the correct cell in sheet2
            sheet2[new_string].value = sheet[get_string].value
            new_col += 1

    # save all of the documents at the end
    doc.save(sheet_name)


# This line should read the name of the excel document you want read, and the name of the
# sheet that needs to be converted - run_excel('excel_document.xlsx', 'sheet_name')
parotid('Contralateral Parotid (test).xlsx', 'Contralateral Parotid')